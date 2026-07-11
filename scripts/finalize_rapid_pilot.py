"""Finalize real rapid-pilot sessions into the Week 6 evaluation package."""

from __future__ import annotations

import csv
import json
import subprocess
import sys
from collections import Counter
from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
RAPID_DIR = ROOT / "study" / "rapid_pilot"
RESPONSES_PATH = RAPID_DIR / "responses.jsonl"
WORKBOOK_PATH = RAPID_DIR / "rapid_pilot_results.xlsx"


def read_sessions() -> list[dict[str, object]]:
    if not RESPONSES_PATH.exists():
        raise SystemExit(f"No rapid pilot responses found at {RESPONSES_PATH}")
    sessions = []
    for line in RESPONSES_PATH.read_text(encoding="utf-8").splitlines():
        if line.strip():
            sessions.append(json.loads(line))
    if len(sessions) < 2:
        raise SystemExit("Collect at least 2 complete participants before finalizing statistics.")
    return sessions


def raw_tlx_mean(ratings: dict[str, object]) -> float:
    fields = ["mental", "physical", "temporal", "performance", "effort", "frustration"]
    return sum(float(ratings[field]) for field in fields) / len(fields)


def interview_codes(text: str) -> tuple[str, str, str, str, str]:
    lowered = text.lower()
    codes: list[str] = []
    if any(word in lowered for word in ["role", "meaning", "understand", "clear", "help"]):
        codes.append("semantic_clarity")
    if any(word in lowered for word in ["guess", "confusing", "hard", "lost", "memory"]):
        codes.append("notation_confusion")
    if any(word in lowered for word in ["mental", "effort", "workload", "remember"]):
        codes.append("cognitive_load_high")
    if any(word in lowered for word in ["symbol", "variable", "matrix", "determinant", "limit", "integral"]):
        codes.append("symbol_role_helpful")
    if any(word in lowered for word in ["fast", "slow", "pace", "pacing", "speed"]):
        codes.append("pacing_issue")
    if any(word in lowered for word in ["semantic", "version b", "second version"]):
        codes.append("audio_preference_semantic")
    if not codes:
        codes.append("domain_familiarity")
    codes = (codes + ["", "", ""])[:3]
    valence = "positive" if any(word in lowered for word in ["better", "help", "clear", "prefer"]) else "mixed"
    condition = "mathontospeak_semantic" if "semantic" in lowered or "version b" in lowered else "both"
    return codes[0], codes[1], codes[2], valence, condition


def write_workbook(sessions: list[dict[str, object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Comprehension Scores"
    ws.append(["Participant ID", "Condition", "Stimulus ID", "Question ID", "Response", "Correct Choice", "Correct", "Score"])
    for session in sessions:
        for row in session["responses"]:
            response = str(row["response"])
            correct_choice = str(row["correct_choice"])
            score = 1 if response == correct_choice else 0
            ws.append(
                [
                    session["participant_id"],
                    row["condition"],
                    row["stimulus_id"],
                    row["question_id"],
                    response,
                    correct_choice,
                    "Yes" if score else "No",
                    score,
                ]
            )

    ws = wb.create_sheet("NASA TLX")
    ws.append(["Participant ID", "Condition", "Mental", "Physical", "Temporal", "Performance", "Effort", "Frustration", "Notes", "Raw TLX Mean"])
    for session in sessions:
        for condition, ratings in session["nasa"].items():
            ws.append(
                [
                    session["participant_id"],
                    condition,
                    ratings["mental"],
                    ratings["physical"],
                    ratings["temporal"],
                    ratings["performance"],
                    ratings["effort"],
                    ratings["frustration"],
                    "Rapid pilot participant entry.",
                    raw_tlx_mean(ratings),
                ]
            )

    ws = wb.create_sheet("Interview Coding")
    ws.append(["Participant ID", "Excerpt / Summary", "Code 1", "Code 2", "Code 3", "Valence", "Condition Mentioned", "Researcher Notes"])
    for session in sessions:
        interview = session["interview"]
        summary = " ".join(str(interview.get(key, "")).strip() for key in ["preference", "version_a", "version_b"]).strip()
        code_1, code_2, code_3, valence, condition = interview_codes(summary)
        ws.append([session["participant_id"], summary, code_1, code_2, code_3, valence, condition, "Collected through rapid pilot runner."])

    ws = wb.create_sheet("Pilot Provenance")
    ws.append(["Field", "Value"])
    ws.append(["Data type", "Real rapid pilot participant entries"])
    ws.append(["Participants", len(sessions)])
    ws.append(["Source", str(RESPONSES_PATH.relative_to(ROOT))])
    ws.append(["Design", "Within-subjects, counterbalanced AB/BA"])

    RAPID_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(WORKBOOK_PATH)


def export_csvs(sessions: list[dict[str, object]]) -> None:
    comprehension_path = RAPID_DIR / "rapid_pilot_comprehension_scores.csv"
    nasa_path = RAPID_DIR / "rapid_pilot_nasa_tlx.csv"
    interview_path = RAPID_DIR / "rapid_pilot_interview_coding.csv"

    with comprehension_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["participant_id", "condition", "stimulus_id", "question_id", "response", "correct_choice", "score"])
        for session in sessions:
            for row in session["responses"]:
                score = 1 if row["response"] == row["correct_choice"] else 0
                writer.writerow([session["participant_id"], row["condition"], row["stimulus_id"], row["question_id"], row["response"], row["correct_choice"], score])

    with nasa_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["participant_id", "condition", "mental", "physical", "temporal", "performance", "effort", "frustration", "raw_tlx_mean"])
        for session in sessions:
            for condition, ratings in session["nasa"].items():
                writer.writerow([session["participant_id"], condition, ratings["mental"], ratings["physical"], ratings["temporal"], ratings["performance"], ratings["effort"], ratings["frustration"], raw_tlx_mean(ratings)])

    with interview_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["participant_id", "summary", "code_1", "code_2", "code_3", "valence", "condition_mentioned"])
        for session in sessions:
            summary = " ".join(str(session["interview"].get(key, "")).strip() for key in ["preference", "version_a", "version_b"]).strip()
            writer.writerow([session["participant_id"], summary, *interview_codes(summary)])


def print_session_summary(sessions: list[dict[str, object]]) -> None:
    counts = Counter()
    for session in sessions:
        counts["participants"] += 1
        counts["mcq_rows"] += len(session["responses"])
        counts["nasa_rows"] += len(session["nasa"])
        counts["interviews"] += 1
    print(f"Participants: {counts['participants']}")
    print(f"MCQ rows: {counts['mcq_rows']}")
    print(f"NASA-TLX rows: {counts['nasa_rows']}")
    print(f"Interview rows: {counts['interviews']}")


def main() -> None:
    sessions = read_sessions()
    write_workbook(sessions)
    export_csvs(sessions)
    print_session_summary(sessions)
    print(f"Saved {WORKBOOK_PATH}")

    cmd = [sys.executable, str(ROOT / "scripts" / "week6_evaluation_package.py"), "--workbook", str(WORKBOOK_PATH.relative_to(ROOT))]
    subprocess.run(cmd, cwd=ROOT, check=True)
    print("Regenerated Week 6 evaluation package from rapid pilot workbook.")


if __name__ == "__main__":
    main()
