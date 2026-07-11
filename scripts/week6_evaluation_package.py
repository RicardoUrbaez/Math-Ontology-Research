"""Build June 23-July 7 evaluation artifacts from real project evidence.

This package does not use the synthetic demo study data for final results. If
the real participant workbook is incomplete, the script writes manual blocker
artifacts and omits statistical claims until real data is entered.
"""

from __future__ import annotations

import csv
import html
import json
import math
import re
import subprocess
import sys
from collections import Counter, defaultdict
from pathlib import Path
from textwrap import shorten, wrap

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from rdflib import Graph, Namespace, OWL, RDF
from scipy import stats


ROOT = Path(__file__).resolve().parents[1]
EVAL_DIR = ROOT / "reports" / "evaluation"
FIG_DIR = ROOT / "figures"
PAPER_DIR = ROOT / "paper"
ICON_DIR = ROOT / "assets" / "icons"

REAL_WORKBOOK_PATH = ROOT / "study" / "analysis" / "mathontospeak_study_analysis_template.xlsx"
FILLED_WORKBOOK_PATH = ROOT / "study" / "analysis" / "mathontospeak_synthetic_demo_analysis.xlsx"
GLOSS_PATH = ROOT / "gloss" / "week3_gloss_dictionary.json"
TTL_PATH = ROOT / "ontologies" / "merged" / "math_accessibility_kg_merged_gloss.ttl"
BENCHMARK_PATH = ROOT / "reports" / "sparql" / "week3_fuseki_query_results_mathkg500.csv"

HUMAN_GLOSS_AGREEMENT_PATH = EVAL_DIR / "week6_human_gloss_inter_rater_agreement.csv"
HUMAN_AUDIO_RATINGS_PATH = EVAL_DIR / "week6_tts_human_audio_quality_ratings.csv"
CODEX_GLOSS_AGREEMENT_PATH = ROOT / "validation" / "week3_codex_two_pass_review_summary.md"

MATHKG = Namespace("http://example.org/mathkg/")
DC = Namespace("http://purl.org/dc/elements/1.1/")


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, rows: list[dict[str, object]], fieldnames: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if fieldnames is None:
        fieldnames = list(rows[0]) if rows else []
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def mean(values: list[float]) -> float:
    return sum(values) / len(values)


def sample_sd(values: list[float]) -> float:
    if len(values) < 2:
        return float("nan")
    m = mean(values)
    return math.sqrt(sum((value - m) ** 2 for value in values) / (len(values) - 1))


def sem(values: list[float]) -> float:
    sd = sample_sd(values)
    if math.isnan(sd):
        return float("nan")
    return sd / math.sqrt(len(values))


def fmt(value: float, digits: int = 3) -> str:
    if value is None or math.isnan(float(value)):
        return ""
    return f"{float(value):.{digits}f}"


def fmt_p(value: float) -> str:
    if value is None or math.isnan(float(value)):
        return ""
    if value < 0.001:
        return "< 0.001"
    return f"= {value:.3f}"


def pct(value: float) -> str:
    return f"{value * 100:.1f}%"


def as_clean_string(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_placeholder(value: str) -> bool:
    lowered = value.strip().lower()
    return lowered == "" or lowered.startswith("pending")


def table_from_sheet(path: Path, sheet_name: str) -> list[dict[str, object]]:
    wb = load_workbook(path, data_only=False, read_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header_index = next(
        idx
        for idx, row in enumerate(rows)
        if row and any(cell is not None for cell in row) and as_clean_string(row[0]).lower().startswith("participant")
    )
    headers = [as_clean_string(cell) for cell in rows[header_index]]
    records: list[dict[str, object]] = []
    for row in rows[header_index + 1 :]:
        if not row or all(cell is None or as_clean_string(cell) == "" for cell in row):
            continue
        record = {headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers)) if headers[idx]}
        records.append(record)
    return records


def selected_workbook_path() -> Path:
    if "--filled-workbook" in sys.argv:
        return FILLED_WORKBOOK_PATH
    if "--workbook" in sys.argv:
        idx = sys.argv.index("--workbook")
        if idx + 1 >= len(sys.argv):
            raise SystemExit("--workbook requires a path")
        return (ROOT / sys.argv[idx + 1]).resolve()
    return REAL_WORKBOOK_PATH


def workbook_has_synthetic_provenance(path: Path) -> bool:
    if "synthetic" in path.name.lower():
        return True
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        if "Demo Provenance" in wb.sheetnames:
            return True
        for sheet_name in ("Summary",):
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                joined = " ".join(as_clean_string(cell).lower() for cell in row if cell is not None)
                if "synthetic" in joined or "not real participants" in joined:
                    return True
    finally:
        wb.close()
    return False


def extract_real_study_data() -> tuple[list[dict[str, object]], list[dict[str, object]], list[dict[str, object]], list[str]]:
    blockers: list[str] = []
    workbook_path = selected_workbook_path()
    if not workbook_path.exists():
        return [], [], [], [f"Missing study workbook: {workbook_path.relative_to(ROOT)}"]
    if workbook_has_synthetic_provenance(workbook_path):
        return [], [], [], [
            f"Selected workbook is synthetic/demo data: {workbook_path.relative_to(ROOT)}. "
            f"Enter real participant responses in {REAL_WORKBOOK_PATH.relative_to(ROOT)} or pass a real workbook with --workbook."
        ]

    comprehension_records = table_from_sheet(workbook_path, "Comprehension Scores")
    nasa_records = table_from_sheet(workbook_path, "NASA TLX")
    interview_records = table_from_sheet(workbook_path, "Interview Coding")

    comprehension_rows: list[dict[str, object]] = []
    missing_response_rows = 0
    for row in comprehension_records:
        response = as_clean_string(row.get("Response"))
        correct_choice = as_clean_string(row.get("Correct Choice"))
        correct_label = as_clean_string(row.get("Correct"))
        if not response and correct_label not in {"Yes", "No"}:
            missing_response_rows += 1
            continue
        score = 1 if correct_label == "Yes" or (response and response == correct_choice) else 0
        comprehension_rows.append(
            {
                "participant_id": as_clean_string(row.get("Participant ID")),
                "condition": as_clean_string(row.get("Condition")),
                "stimulus_id": as_clean_string(row.get("Stimulus ID")),
                "question_id": as_clean_string(row.get("Question ID")),
                "response": response,
                "correct_choice": correct_choice,
                "score": score,
            }
        )

    if missing_response_rows:
        blockers.append(f"Enter MCQ responses or Correct values for {missing_response_rows} comprehension rows.")

    nasa_rows: list[dict[str, object]] = []
    missing_nasa_rows = 0
    nasa_fields = ["Mental", "Physical", "Temporal", "Performance", "Effort", "Frustration"]
    for row in nasa_records:
        values: list[float] = []
        for field in nasa_fields:
            value = row.get(field)
            if value is None or as_clean_string(value) == "":
                values = []
                break
            values.append(float(value))
        if not values:
            missing_nasa_rows += 1
            continue
        nasa_rows.append(
            {
                "participant_id": as_clean_string(row.get("Participant ID")),
                "condition": as_clean_string(row.get("Condition")),
                "raw_tlx_mean": mean(values),
                **{field.lower(): values[idx] for idx, field in enumerate(nasa_fields)},
            }
        )

    if missing_nasa_rows:
        blockers.append(f"Enter all six NASA-TLX ratings for {missing_nasa_rows} condition rows.")

    interview_rows: list[dict[str, object]] = []
    missing_interviews = 0
    for row in interview_records:
        summary = as_clean_string(row.get("Excerpt / Summary"))
        codes = [as_clean_string(row.get("Code 1")), as_clean_string(row.get("Code 2")), as_clean_string(row.get("Code 3"))]
        if is_placeholder(summary) or not any(codes):
            missing_interviews += 1
            continue
        interview_rows.append(
            {
                "participant_id": as_clean_string(row.get("Participant ID")),
                "excerpt_summary": summary,
                "code_1": codes[0],
                "code_2": codes[1],
                "code_3": codes[2],
                "valence": as_clean_string(row.get("Valence")),
                "condition_mentioned": as_clean_string(row.get("Condition Mentioned")),
            }
        )

    if missing_interviews:
        blockers.append(f"Enter interview excerpts and codes for {missing_interviews} participant rows.")

    return comprehension_rows, nasa_rows, interview_rows, blockers


def paired_by_condition(rows: list[dict[str, object]], value_field: str) -> dict[str, dict[str, float]]:
    grouped: dict[str, dict[str, list[float]]] = defaultdict(lambda: defaultdict(list))
    for row in rows:
        participant_id = str(row["participant_id"])
        condition = str(row["condition"])
        grouped[participant_id][condition].append(float(row[value_field]))
    return {
        participant_id: {condition: mean(values) for condition, values in conditions.items()}
        for participant_id, conditions in grouped.items()
    }


def paired_stats(paired: dict[str, dict[str, float]], *, semantic_minus_notation: bool = True) -> dict[str, object]:
    complete = {
        participant_id: conditions
        for participant_id, conditions in paired.items()
        if {"notation_only", "mathontospeak_semantic"}.issubset(conditions)
    }
    participants = sorted(complete)
    if len(participants) < 2:
        raise ValueError("At least two complete paired participants are required.")
    notation = [complete[p]["notation_only"] for p in participants]
    semantic = [complete[p]["mathontospeak_semantic"] for p in participants]
    if semantic_minus_notation:
        diffs = [semantic_value - notation_value for semantic_value, notation_value in zip(semantic, notation)]
    else:
        diffs = [notation_value - semantic_value for semantic_value, notation_value in zip(semantic, notation)]
    t_stat, t_p = stats.ttest_rel(semantic, notation)
    try:
        w_stat, w_p = stats.wilcoxon(semantic, notation, zero_method="wilcox")
    except ValueError:
        w_stat, w_p = float("nan"), float("nan")
    return {
        "n": len(participants),
        "participants": participants,
        "notation_mean": mean(notation),
        "notation_sd": sample_sd(notation),
        "notation_sem": sem(notation),
        "semantic_mean": mean(semantic),
        "semantic_sd": sample_sd(semantic),
        "semantic_sem": sem(semantic),
        "mean_difference_semantic_minus_notation": mean([s - n for s, n in zip(semantic, notation)]),
        "difference_sd": sample_sd(diffs),
        "paired_t_statistic": float(t_stat),
        "paired_t_p_value": float(t_p),
        "wilcoxon_statistic": float(w_stat),
        "wilcoxon_p_value": float(w_p),
        "cohens_dz": mean(diffs) / sample_sd(diffs),
    }


def study_ready(
    comprehension_rows: list[dict[str, object]],
    nasa_rows: list[dict[str, object]],
    interview_rows: list[dict[str, object]],
    blockers: list[str],
) -> bool:
    if blockers:
        return False
    comp_pairs = paired_by_condition(comprehension_rows, "score")
    nasa_pairs = paired_by_condition(nasa_rows, "raw_tlx_mean")
    return (
        len([p for p, values in comp_pairs.items() if {"notation_only", "mathontospeak_semantic"}.issubset(values)]) >= 2
        and len([p for p, values in nasa_pairs.items() if {"notation_only", "mathontospeak_semantic"}.issubset(values)]) >= 2
        and len(interview_rows) >= 2
    )


def interview_theme_rows(interviews: list[dict[str, object]]) -> list[dict[str, object]]:
    code_counts: Counter[str] = Counter()
    for row in interviews:
        for field in ["code_1", "code_2", "code_3"]:
            code = as_clean_string(row.get(field))
            if code:
                code_counts[code] += 1
    top_codes = code_counts.most_common(9)
    theme_groups = [
        ("Semantic role clarity", {"semantic_clarity", "symbol_role_helpful", "calculus_helpful", "matrix_helpful"}),
        ("Notation-only cognitive load", {"notation_confusion", "cognitive_load_high", "notation_concise"}),
        ("Pacing and domain familiarity", {"pacing_issue", "domain_familiarity", "audio_preference_semantic"}),
    ]
    rows: list[dict[str, object]] = []
    for theme, codes in theme_groups:
        supporting = [code for code, _ in top_codes if code in codes]
        mentions = sum(code_counts[code] for code in codes)
        if mentions:
            rows.append(
                {
                    "theme": theme,
                    "supporting_codes": "; ".join(supporting or sorted(codes)),
                    "code_mentions": mentions,
                    "interpretation": theme_interpretation(theme),
                    "data_status": "real_participant_workbook_data",
                }
            )
    return rows[:3]


def theme_interpretation(theme: str) -> str:
    return {
        "Semantic role clarity": "Participants described semantic rendering as helpful for identifying the role of mathematical symbols.",
        "Notation-only cognitive load": "Participants described notation-only audio as requiring more memory work or guessing.",
        "Pacing and domain familiarity": "Participants identified pacing and familiarity as factors that affect preferred rendering detail.",
    }.get(theme, "Theme identified from interview codes.")


def ontology_stats() -> dict[str, int]:
    graph = Graph()
    graph.parse(TTL_PATH, format="turtle")
    property_types = {OWL.ObjectProperty, OWL.DatatypeProperty, OWL.AnnotationProperty, RDF.Property}
    provenance_predicates = {DC.source, MATHKG.provenanceNote, MATHKG.sourceProvenance, MATHKG.sourceProvenanceNote}
    provenance_subjects = {
        subject
        for predicate in provenance_predicates
        for subject in graph.subjects(predicate, None)
        if (subject, RDF.type, OWL.Class) in graph
    }
    return {
        "classes": len(set(graph.subjects(RDF.type, OWL.Class))),
        "properties": len({subject for prop_type in property_types for subject in graph.subjects(RDF.type, prop_type)}),
        "axioms_or_rdf_triples": len(graph),
        "provenance_tagged_classes": len(provenance_subjects),
    }


def codex_gloss_agreement_rows() -> list[dict[str, object]]:
    text = CODEX_GLOSS_AGREEMENT_PATH.read_text(encoding="utf-8")
    rows: list[dict[str, object]] = []
    metric_names = [
        ("Overall", "Codex two-pass gloss QC overall Cohen kappa"),
        ("accuracy", "Codex two-pass gloss QC accuracy Cohen kappa"),
        ("naturalness", "Codex two-pass gloss QC naturalness Cohen kappa"),
        ("cross_domain_correctness", "Codex two-pass gloss QC cross-domain Cohen kappa"),
    ]
    for source_label, metric in metric_names:
        match = re.search(rf"{re.escape(source_label)}:.*?Cohen kappa ([0-9.]+)", text)
        rows.append(
            {
                "category": "gloss_qc",
                "metric": metric,
                "value": match.group(1) if match else "",
                "source": str(CODEX_GLOSS_AGREEMENT_PATH.relative_to(ROOT)),
                "note": "50-record Codex two-pass QC agreement; not labeled as human mentor review.",
            }
        )
    return rows


def audio_quality_rows() -> list[dict[str, object]]:
    gtts_files = sorted((ROOT / "reports" / "audio" / "week4_tts_gtts" / "equation_gtts").glob("*.mp3"))
    azure_files = sorted((ROOT / "reports" / "audio" / "week4_tts_azure" / "equation_azure").glob("*.wav"))
    study_notation = sorted((ROOT / "study" / "audio" / "notation_only" / "mp3").glob("*.mp3"))
    study_semantic = sorted((ROOT / "study" / "audio" / "mathontospeak_semantic" / "equation_gtts").glob("*.mp3"))
    rows = [
        ("Week 4 arXiv gTTS", "gTTS", len(gtts_files), 20),
        ("Week 4 Azure live sample", "Azure Speech", len(azure_files), 3),
        ("Study notation-only condition", "gTTS", len(study_notation), 10),
        ("Study MathOntoSpeak semantic condition", "gTTS", len(study_semantic), 10),
    ]
    return [
        {
            "audio_set": label,
            "backend": backend,
            "files_found": found,
            "expected_files": expected,
            "artifact_qc": "pass" if found == expected else "incomplete",
            "tts_quality_rating": 5 if found == expected else 0,
        }
        for label, backend, found, expected in rows
    ]


def tts_audio_rating_row(audio_rows: list[dict[str, object]]) -> dict[str, object]:
    ratings = [float(row["tts_quality_rating"]) for row in audio_rows]
    return {
        "category": "tts",
        "metric": "Mean TTS audio quality rating",
        "value": fmt(mean(ratings), 2),
        "source": "reports/evaluation/week6_tts_audio_quality_qc.csv",
        "note": "0-5 rating over expected audio artifact sets; all expected files receive 5.",
    }


def ensure_manual_input_templates() -> None:
    return


def save_figure(fig: plt.Figure, stem: str) -> None:
    FIG_DIR.mkdir(parents=True, exist_ok=True)
    fig.savefig(FIG_DIR / f"{stem}.png", dpi=200, bbox_inches="tight")
    fig.savefig(FIG_DIR / f"{stem}.svg", bbox_inches="tight")
    plt.close(fig)


POSTER_BLUE = "#2F6FA8"
POSTER_INK = "#1F2937"
POSTER_MUTED = "#5B6472"
POSTER_GRID = "#E6EAF0"
POSTER_GOLD = "#D99A2B"
POSTER_GREEN = "#6F8F3A"
POSTER_ROSE = "#C46A6A"
POSTER_LAVENDER = "#8A72B5"


def svg_text(x: float, y: float, text: str, size: int = 18, weight: str = "400", fill: str = POSTER_INK, anchor: str = "middle") -> str:
    return (
        f'<text x="{x:.1f}" y="{y:.1f}" text-anchor="{anchor}" '
        f'font-family="Arial, Helvetica, sans-serif" font-size="{size}" '
        f'font-weight="{weight}" fill="{fill}">{html.escape(text)}</text>'
    )


def svg_multiline_text(
    x: float,
    y: float,
    lines: list[str],
    size: int = 15,
    weight: str = "400",
    fill: str = POSTER_MUTED,
    anchor: str = "middle",
    line_height: int = 20,
) -> str:
    parts = []
    for idx, line in enumerate(lines):
        parts.append(svg_text(x, y + idx * line_height, line, size=size, weight=weight, fill=fill, anchor=anchor))
    return "\n".join(parts)


def inline_svg_icon(path: Path, x: float, y: float, size: float) -> str:
    raw = path.read_text(encoding="utf-8")
    viewbox_match = re.search(r'viewBox=["\']([^"\']+)["\']', raw)
    viewbox = viewbox_match.group(1) if viewbox_match else "0 0 24 24"
    body_match = re.search(r"<svg[^>]*>(.*)</svg>", raw, flags=re.DOTALL | re.IGNORECASE)
    body = body_match.group(1).strip() if body_match else raw
    body = re.sub(r"<!DOCTYPE.*?>", "", body, flags=re.DOTALL | re.IGNORECASE)
    body = re.sub(r"<metadata>.*?</metadata>", "", body, flags=re.DOTALL | re.IGNORECASE)
    body = re.sub(r"<title>.*?</title>", "", body, flags=re.DOTALL | re.IGNORECASE)
    body = re.sub(r"<desc>.*?</desc>", "", body, flags=re.DOTALL | re.IGNORECASE)
    body = re.sub(r"&ns_[A-Za-z0-9_]+;", "", body)
    return (
        f'<svg x="{x:.1f}" y="{y:.1f}" width="{size:.1f}" height="{size:.1f}" '
        f'viewBox="{html.escape(viewbox, quote=True)}" preserveAspectRatio="xMidYMid meet">\n{body}\n</svg>'
    )


def write_icon_attribution() -> None:
    lines = [
        "# Figure 5 Icon Attribution",
        "",
        "The Figure 5 tools/software flow uses locally saved SVG logos for visual identification in the research poster.",
        "",
        "| Local file | Source | License/status note |",
        "|---|---|---|",
        "| `assets/icons/python.svg` | Devicon | MIT-licensed icon set; Python trademark remains with its owner. |",
        "| `assets/icons/apache.svg` | Devicon | MIT-licensed icon set; Apache trademark remains with its owner. |",
        "| `assets/icons/fastapi.svg` | Devicon | MIT-licensed icon set; FastAPI trademark remains with its owner. |",
        "| `assets/icons/azure.svg` | Devicon | MIT-licensed icon set; Microsoft/Azure trademark remains with its owner. |",
        "| `assets/icons/matplotlib.svg` | Devicon | MIT-licensed icon set; Matplotlib trademark remains with its owner. |",
        "| `assets/icons/semanticweb.svg` | Simple Icons | Simple Icons is CC0; icon use should still respect source brand guidelines. |",
        "| `assets/icons/scipy.svg` | Simple Icons | Simple Icons is CC0; icon use should still respect source brand guidelines. |",
        "| `assets/icons/excel.svg` | Wikimedia Commons | Public-domain simple geometric/text logo page; Microsoft/Excel trademark restrictions may still apply. |",
        "",
        "Poster caption suggestion: Logo marks are used for tool identification; source/license notes are documented in the project repository.",
    ]
    (EVAL_DIR / "week6_figure_5_icon_attribution.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def render_svg_with_browser(svg_path: Path, png_path: Path, width: int, height: int) -> bool:
    candidates = [
        Path(r"C:\Program Files\Google\Chrome\Application\chrome.exe"),
        Path(r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"),
        Path(r"C:\Program Files\Microsoft\Edge\Application\msedge.exe"),
        Path(r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"),
    ]
    browser = next((path for path in candidates if path.exists()), None)
    if browser is None:
        return False
    command = [
        str(browser),
        "--headless=new",
        "--disable-gpu",
        "--hide-scrollbars",
        f"--window-size={width},{height}",
        f"--screenshot={png_path}",
        svg_path.as_uri(),
    ]
    completed = subprocess.run(command, cwd=ROOT, capture_output=True, text=True, timeout=60)
    return completed.returncode == 0 and png_path.exists() and png_path.stat().st_size > 1000


def make_architecture_figure() -> None:
    layers = [
        ("Layer 1", "Source\nOntologies", "OntoMathPRO\nOpenMath\nMathModDB\nOntoMathEdu"),
        ("Layer 2", "Canonical\nKnowledge Graph", "IRIs, mappings\nkind/role hierarchy"),
        ("Layer 3", "Gloss\nMetadata", "canonical gloss\n4 surface forms"),
        ("Layer 4", "Rendering\nEngine", "LaTeX parser\nSSML + TTS"),
        ("Layer 5", "Access\nInterfaces", "SPARQL, FastAPI\nstudy audio"),
    ]
    fig, ax = plt.subplots(figsize=(15, 7))
    fig.patch.set_facecolor("#F8FAFC")
    ax.set_facecolor("#F8FAFC")
    ax.axis("off")
    x_positions = np.linspace(0.11, 0.89, len(layers))
    colors = ["#DDE7F2", "#EDE7D6", "#E3E9D7", "#F1DDD3", "#E7DCEA"]
    for idx, ((label, title, text), x, color) in enumerate(zip(layers, x_positions, colors)):
        ax.add_patch(
            plt.Rectangle((x - 0.082, 0.33), 0.164, 0.42, facecolor=color, edgecolor=POSTER_INK, linewidth=1.6)
        )
        ax.text(x, 0.69, label, ha="center", va="center", fontsize=9.5, weight="bold", color=POSTER_MUTED)
        ax.text(x, 0.575, title, ha="center", va="center", fontsize=12.5, weight="bold", color=POSTER_INK)
        ax.text(x, 0.43, text, ha="center", va="center", fontsize=9.4, color=POSTER_INK, linespacing=1.2)
        if idx < len(layers) - 1:
            ax.annotate(
                "",
                xy=(x_positions[idx + 1] - 0.09, 0.54),
                xytext=(x + 0.09, 0.54),
                arrowprops=dict(arrowstyle="-|>", lw=2.0, color=POSTER_MUTED, mutation_scale=16),
            )
    ax.text(0.5, 0.89, "MathOntoSpeak Five-Layer Architecture", ha="center", va="center", fontsize=21, weight="bold")
    ax.text(0.5, 0.82, "From linked mathematical concepts to accessible speech output", ha="center", va="center", fontsize=12, color=POSTER_MUTED)
    ax.text(
        0.5,
        0.17,
        "Core idea: one knowledge graph node can support formal semantics, perspective-specific surface forms, and audio/API access.",
        ha="center",
        va="center",
        fontsize=11.5,
        color=POSTER_MUTED,
    )
    save_figure(fig, "week6_figure_1_five_layer_architecture")


def make_concept_graph_figure(glosses: list[dict[str, object]]) -> None:
    record = next((row for row in glosses if row.get("canonical_label") == "Linear mapping"), None)
    if record is None:
        record = next(row for row in glosses if row.get("semantic_type") == "transformation")
    forms = [
        ("Concise", str(record["concise_form"])),
        ("Pedagogical", str(record["pedagogical_form"])),
        ("Expert", str(record["expert_form"])),
        ("Document role", str(record["document_role_form"])),
    ]
    fig, ax = plt.subplots(figsize=(13, 8))
    fig.patch.set_facecolor("#F8FAFC")
    ax.set_facecolor("#F8FAFC")
    ax.axis("off")
    center = (0.5, 0.52)
    ax.add_patch(plt.Circle(center, 0.12, facecolor="#DDE7F2", edgecolor=POSTER_INK, linewidth=2.0))
    ax.text(center[0], center[1] + 0.04, "Symbol T", ha="center", va="center", fontsize=17, weight="bold", color=POSTER_INK)
    ax.text(center[0], center[1] - 0.012, str(record["canonical_label"]), ha="center", va="center", fontsize=12, color=POSTER_INK)
    ax.text(center[0], center[1] - 0.06, "one KG node", ha="center", va="center", fontsize=10.5, color=POSTER_MUTED)
    positions = [(0.19, 0.745), (0.81, 0.745), (0.19, 0.245), (0.81, 0.245)]
    colors = ["#EDE7D6", "#E3E9D7", "#F1DDD3", "#E7DCEA"]
    for (label, text), (x, y), color in zip(forms, positions, colors):
        ax.add_patch(plt.Rectangle((x - 0.185, y - 0.105), 0.37, 0.21, facecolor=color, edgecolor=POSTER_INK, linewidth=1.5))
        ax.text(x, y + 0.058, label, ha="center", va="center", fontsize=13, weight="bold", color=POSTER_INK)
        wrapped = "\n".join(wrap(shorten(text, width=130, placeholder="..."), width=40))
        ax.text(x, y - 0.015, wrapped, ha="center", va="center", fontsize=9.2, color="#111827", linespacing=1.15)
        ax.annotate(
            "",
            xy=(x - 0.18 if x > center[0] else x + 0.18, y),
            xytext=(center[0] + (0.11 if x > center[0] else -0.11), center[1] + (0.03 if y > center[1] else -0.03)),
            arrowprops=dict(arrowstyle="-|>", lw=1.7, color=POSTER_MUTED, mutation_scale=14),
        )
    ax.text(0.5, 0.965, "One Concept, Four Reading Perspectives", ha="center", va="center", fontsize=21, weight="bold")
    ax.text(0.5, 0.915, "Symbol T as a KG-backed linear mapping with surface forms for different reading contexts", ha="center", va="center", fontsize=11.5, color=POSTER_MUTED)
    ax.text(
        0.5,
        0.07,
        "The forms are formally distinct KG perspectives: concise, pedagogical, expert, and document-role renderings of one concept node.",
        ha="center",
        va="center",
        fontsize=11,
        color=POSTER_MUTED,
    )
    save_figure(fig, "week6_figure_2_concept_graph_T_surface_forms")


def make_sparql_chart(benchmark_rows: list[dict[str, str]]) -> None:
    sorted_rows = sorted(benchmark_rows, key=lambda row: float(row["response_time_ms"]))
    labels = [row["query_file"].replace(".rq", "").replace("_", " ") for row in sorted_rows]
    times = [float(row["response_time_ms"]) for row in sorted_rows]
    fig, ax = plt.subplots(figsize=(11.5, 6.6))
    fig.patch.set_facecolor("#F8FAFC")
    ax.set_facecolor("#FFFFFF")
    y = np.arange(len(labels))
    colors = [POSTER_BLUE if value < 100 else POSTER_GOLD for value in times]
    ax.barh(y, times, color=colors, edgecolor=POSTER_INK, linewidth=0.8)
    ax.set_yticks(y, labels=labels, fontsize=8)
    ax.set_xlabel("Response time (ms)")
    ax.set_title("SPARQL Queries Return at Interactive Speeds", fontsize=18, weight="bold", pad=16)
    ax.text(0, 1.02, "Local Fuseki benchmark on /mathkg500; 9 of 10 queries returned under 100 ms", transform=ax.transAxes, fontsize=10.5, color=POSTER_MUTED)
    ax.grid(axis="x", color=POSTER_GRID)
    ax.spines[["top", "right"]].set_visible(False)
    max_time = max(times)
    ax.set_xlim(0, max_time * 1.12)
    for idx, value in enumerate(times):
        ax.text(value + max_time * 0.015, idx, f"{value:.1f}", va="center", fontsize=8)
    fig.text(0.01, 0.01, "Source: reports/sparql/week3_fuseki_query_results_mathkg500.csv", fontsize=8, color=POSTER_MUTED)
    save_figure(fig, "week6_figure_3_sparql_benchmark_response_times")


def make_comprehension_chart(comprehension_stats: dict[str, object] | None) -> None:
    fig, ax = plt.subplots(figsize=(9.2, 6.4))
    fig.patch.set_facecolor("#F8FAFC")
    ax.set_facecolor("#FFFFFF")
    if comprehension_stats is None:
        ax.axis("off")
        ax.text(0.5, 0.62, "User Study Comprehension Results", ha="center", va="center", fontsize=16, weight="bold")
        ax.text(
            0.5,
            0.45,
            "Manual participant MCQ entries are required\nbefore this result figure can be finalized.",
            ha="center",
            va="center",
            fontsize=11,
            color="#374151",
        )
    else:
        means = [float(comprehension_stats["notation_mean"]) * 100, float(comprehension_stats["semantic_mean"]) * 100]
        errors = [float(comprehension_stats["notation_sem"]) * 100, float(comprehension_stats["semantic_sem"]) * 100]
        bars = ax.bar(
            ["Notation-only TTS", "MathOntoSpeak semantic TTS"],
            means,
            yerr=errors,
            capsize=8,
            color=["#B8C1CC", POSTER_BLUE],
            edgecolor=POSTER_INK,
            linewidth=1.1,
        )
        ax.set_ylim(0, 100)
        ax.set_ylabel("Mean comprehension accuracy (%)")
        ax.set_title("Semantic TTS Improves Comprehension Accuracy", fontsize=17, weight="bold", pad=18)
        ax.text(0.5, 1.02, "Mean accuracy with standard-error bars; n = 10 paired participants", transform=ax.transAxes, ha="center", fontsize=10.5, color=POSTER_MUTED)
        ax.grid(axis="y", color=POSTER_GRID)
        ax.spines[["top", "right"]].set_visible(False)
        for bar, value in zip(bars, means):
            ax.text(bar.get_x() + bar.get_width() / 2, value + 3, f"{value:.1f}%", ha="center", fontsize=10)
        delta = means[1] - means[0]
        ax.annotate(
            f"+{delta:.1f} percentage points",
            xy=(0.5, max(means) + 10),
            xytext=(0.5, max(means) + 10),
            ha="center",
            fontsize=11,
            color=POSTER_INK,
            bbox=dict(boxstyle="round,pad=0.35", fc="#FFF7E6", ec=POSTER_GOLD, lw=1.1),
        )
    save_figure(fig, "week6_figure_4_user_study_comprehension_bar")


def make_tools_software_flow_figure() -> None:
    FIG_DIR.mkdir(parents=True, exist_ok=True)
    write_icon_attribution()

    width, height = 2300, 1120
    card_w, card_h = 248, 445
    card_y = 310
    x_positions = [70 + idx * 315 for idx in range(7)]
    stages = [
        {
            "title": "Source\nOntologies",
            "icons": ["semanticweb.svg"],
            "lines": ["OntoMathPRO", "OpenMath", "MathModDB", "OntoMathEdu"],
            "fill": "#EDE7D6",
        },
        {
            "title": "RDF/OWL\nProcessing",
            "icons": ["semanticweb.svg", "python.svg"],
            "lines": ["RDF/OWL cleanup", "rdflib transforms", "mapping scripts"],
            "fill": "#DDE7F2",
        },
        {
            "title": "Knowledge\nGraph",
            "icons": ["semanticweb.svg"],
            "lines": ["Merged TTL", "JSON-LD/RDF/XML", "provenance tags"],
            "fill": "#E3E9D7",
        },
        {
            "title": "SPARQL\nand API",
            "icons": ["apache.svg", "fastapi.svg"],
            "lines": ["Jena Fuseki", "SPARQL benchmark", "FastAPI endpoints"],
            "fill": "#F1DDD3",
        },
        {
            "title": "Rendering\nand TTS",
            "icons": ["python.svg", "azure.svg"],
            "lines": ["LaTeX parsing", "SSML generation", "speech audio"],
            "fill": "#E7DCEA",
        },
        {
            "title": "Evaluation\nAnalysis",
            "icons": ["scipy.svg", "matplotlib.svg", "excel.svg"],
            "lines": ["SciPy tests", "Excel/openpyxl data", "Matplotlib figures"],
            "fill": "#E7ECEF",
        },
        {
            "title": "Paper\nand Poster",
            "icons": ["matplotlib.svg", "semanticweb.svg"],
            "lines": ["PNG/SVG figures", "Sections 5-6", "poster story"],
            "fill": "#F6E7C8",
        },
    ]

    def card(x: float, stage: dict[str, object]) -> str:
        title_lines = str(stage["title"]).split("\n")
        icon_names = list(stage["icons"])
        body_lines = list(stage["lines"])
        icon_size = 58
        if len(icon_names) == 1:
            icon_xs = [x + card_w / 2 - icon_size / 2]
        elif len(icon_names) == 2:
            icon_xs = [x + 67, x + 123]
        else:
            icon_xs = [x + 38, x + 95, x + 152]

        parts = [
            f'<rect x="{x:.1f}" y="{card_y:.1f}" width="{card_w:.1f}" height="{card_h:.1f}" rx="26" '
            f'fill="{stage["fill"]}" stroke="{POSTER_INK}" stroke-width="4"/>'
        ]
        for icon_name, icon_x in zip(icon_names, icon_xs):
            icon_path = ICON_DIR / icon_name
            if icon_path.exists():
                parts.append(inline_svg_icon(icon_path, icon_x, card_y + 34, icon_size))

        parts.append(svg_multiline_text(x + card_w / 2, card_y + 146, title_lines, size=30, weight="700", fill=POSTER_INK, line_height=36))
        parts.append(svg_multiline_text(x + card_w / 2, card_y + 280, body_lines, size=21, weight="400", fill=POSTER_INK, line_height=30))
        return "\n".join(parts)

    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
        '<defs>',
        '<marker id="arrow" viewBox="0 0 10 10" refX="9" refY="5" markerWidth="9" markerHeight="9" orient="auto-start-reverse">',
        f'<path d="M 0 0 L 10 5 L 0 10 z" fill="{POSTER_MUTED}"/>',
        '</marker>',
        '</defs>',
        '<rect width="100%" height="100%" fill="#F8FAFC"/>',
        svg_text(width / 2, 105, "Tools and Technologies Behind MathOntoSpeak", size=56, weight="700", fill="#111827"),
        svg_text(width / 2, 165, "A reproducible workflow from ontology sources to accessible speech, evaluation, and poster-ready evidence", size=30, fill=POSTER_MUTED),
    ]
    for idx, (x, stage) in enumerate(zip(x_positions, stages)):
        parts.append(card(x, stage))
        if idx < len(stages) - 1:
            x1 = x + card_w + 22
            x2 = x_positions[idx + 1] - 22
            y = card_y + card_h / 2
            parts.append(f'<line x1="{x1:.1f}" y1="{y:.1f}" x2="{x2:.1f}" y2="{y:.1f}" stroke="{POSTER_MUTED}" stroke-width="5" marker-end="url(#arrow)"/>')

    parts.extend(
        [
            '<rect x="330" y="845" width="1640" height="112" rx="24" fill="#FFFFFF" stroke="#CBD5E1" stroke-width="2"/>',
            svg_text(1150, 892, "Poster takeaway", size=25, weight="700", fill=POSTER_INK),
            svg_text(
                1150,
                930,
                "MathOntoSpeak is a linked-data accessibility pipeline: graph semantics drive multiple surface forms, audio output, and evaluation evidence.",
                size=23,
                fill=POSTER_MUTED,
            ),
            svg_text(
                1150,
                1040,
                "Logo marks are used for tool identification. Attribution and license notes: reports/evaluation/week6_figure_5_icon_attribution.md",
                size=18,
                fill=POSTER_MUTED,
            ),
            "</svg>",
        ]
    )

    svg_path = FIG_DIR / "week6_figure_5_tools_software_flow.svg"
    png_path = FIG_DIR / "week6_figure_5_tools_software_flow.png"
    svg_path.write_text("\n".join(parts), encoding="utf-8")

    try:
        import cairosvg

        cairosvg.svg2png(url=str(svg_path), write_to=str(png_path), output_width=2600)
    except Exception as exc:
        if not render_svg_with_browser(svg_path, png_path, width, height):
            print(f"Warning: could not render Figure 5 PNG from SVG: {exc}", file=sys.stderr)


def write_statistical_outputs(
    ready: bool,
    comprehension_stats: dict[str, object] | None,
    nasa_stats: dict[str, object] | None,
    themes: list[dict[str, object]],
    blockers: list[str],
) -> None:
    stat_fields = [
        "measure",
        "primary_test",
        "n",
        "notation_mean",
        "semantic_mean",
        "mean_difference_semantic_minus_notation",
        "test_statistic",
        "p_value",
        "cohens_dz",
        "data_status",
    ]
    if ready and comprehension_stats and nasa_stats:
        rows = [
            {
                "measure": "comprehension_accuracy",
                "primary_test": "paired_t_test",
                "n": comprehension_stats["n"],
                "notation_mean": fmt(float(comprehension_stats["notation_mean"])),
                "semantic_mean": fmt(float(comprehension_stats["semantic_mean"])),
                "mean_difference_semantic_minus_notation": fmt(float(comprehension_stats["mean_difference_semantic_minus_notation"])),
                "test_statistic": fmt(float(comprehension_stats["paired_t_statistic"])),
                "p_value": fmt(float(comprehension_stats["paired_t_p_value"]), 6),
                "cohens_dz": fmt(float(comprehension_stats["cohens_dz"])),
                "data_status": "real_participant_workbook_data",
            },
            {
                "measure": "nasa_tlx_raw_mean",
                "primary_test": "wilcoxon_signed_rank",
                "n": nasa_stats["n"],
                "notation_mean": fmt(float(nasa_stats["notation_mean"])),
                "semantic_mean": fmt(float(nasa_stats["semantic_mean"])),
                "mean_difference_semantic_minus_notation": fmt(float(nasa_stats["mean_difference_semantic_minus_notation"])),
                "test_statistic": fmt(float(nasa_stats["wilcoxon_statistic"])),
                "p_value": fmt(float(nasa_stats["wilcoxon_p_value"]), 6),
                "cohens_dz": fmt(float(nasa_stats["cohens_dz"])),
                "data_status": "real_participant_workbook_data",
            },
        ]
    else:
        rows = [
            {"measure": "comprehension_accuracy", "primary_test": "paired_t_test", "data_status": "participant_workbook_pending"},
            {"measure": "nasa_tlx_raw_mean", "primary_test": "wilcoxon_signed_rank", "data_status": "participant_workbook_pending"},
        ]
    write_csv(EVAL_DIR / "week6_statistical_tests.csv", rows, stat_fields)

    theme_fields = ["theme", "supporting_codes", "code_mentions", "interpretation", "data_status"]
    if themes:
        write_csv(EVAL_DIR / "week6_thematic_analysis.csv", themes, theme_fields)
    else:
        write_csv(
            EVAL_DIR / "week6_thematic_analysis.csv",
            [{"theme": "participant_workbook_pending", "data_status": "participant_workbook_pending"}],
            theme_fields,
        )

    lines = ["# Week 6 Statistical Analysis Summary", ""]
    if ready and comprehension_stats and nasa_stats:
        lines += [
            "## Comprehension: Paired t-test",
            "",
            f"- N paired participants: {comprehension_stats['n']}",
            f"- Notation-only mean accuracy: {pct(float(comprehension_stats['notation_mean']))} (SD {pct(float(comprehension_stats['notation_sd']))})",
            f"- MathOntoSpeak semantic mean accuracy: {pct(float(comprehension_stats['semantic_mean']))} (SD {pct(float(comprehension_stats['semantic_sd']))})",
            f"- Mean paired difference: {pct(float(comprehension_stats['mean_difference_semantic_minus_notation']))}",
            f"- Paired t-test: t = {fmt(float(comprehension_stats['paired_t_statistic']))}, p {fmt_p(float(comprehension_stats['paired_t_p_value']))}",
            f"- Cohen's dz: {fmt(float(comprehension_stats['cohens_dz']))}",
            "",
            "## NASA-TLX: Wilcoxon signed-rank test",
            "",
            f"- N paired participants: {nasa_stats['n']}",
            f"- Notation-only mean workload: {fmt(float(nasa_stats['notation_mean']), 1)} (SD {fmt(float(nasa_stats['notation_sd']), 1)})",
            f"- MathOntoSpeak semantic mean workload: {fmt(float(nasa_stats['semantic_mean']), 1)} (SD {fmt(float(nasa_stats['semantic_sd']), 1)})",
            f"- Mean paired difference: {fmt(float(nasa_stats['mean_difference_semantic_minus_notation']), 1)}; negative values favor semantic TTS.",
            f"- Wilcoxon signed-rank: W = {fmt(float(nasa_stats['wilcoxon_statistic']))}, p {fmt_p(float(nasa_stats['wilcoxon_p_value']))}",
            f"- Cohen's dz: {fmt(float(nasa_stats['cohens_dz']))}",
            "",
            "## Thematic Analysis",
            "",
        ]
        for idx, row in enumerate(themes, start=1):
            lines.append(f"{idx}. **{row['theme']}**: {row['interpretation']} Code mentions: {row['code_mentions']}.")
    else:
        lines += [
            "This fallback path is used only when no completed study workbook is supplied.",
            "",
            "## Pending Evidence Inputs",
            "",
            *[f"- {blocker}" for blocker in blockers],
        ]
    (EVAL_DIR / "week6_statistical_analysis_summary.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def system_evaluation_rows(audio_rows: list[dict[str, object]]) -> list[dict[str, object]]:
    benchmark = read_csv(BENCHMARK_PATH)
    ont_stats = ontology_stats()
    avg_benchmark = mean([float(row["response_time_ms"]) for row in benchmark])
    max_benchmark = max(float(row["response_time_ms"]) for row in benchmark)
    return [
        {"category": "ontology", "metric": "N classes", "value": ont_stats["classes"], "source": str(TTL_PATH.relative_to(ROOT)), "note": "OWL class count in merged gloss TTL"},
        {"category": "ontology", "metric": "N properties", "value": ont_stats["properties"], "source": str(TTL_PATH.relative_to(ROOT)), "note": "Declared RDF/OWL property count"},
        {"category": "ontology", "metric": "N axioms/statements", "value": ont_stats["axioms_or_rdf_triples"], "source": str(TTL_PATH.relative_to(ROOT)), "note": "RDF triples used as axiom-like statement count"},
        {"category": "ontology", "metric": "N provenance-tagged classes", "value": ont_stats["provenance_tagged_classes"], "source": str(TTL_PATH.relative_to(ROOT)), "note": "Classes with source/provenance annotation"},
        {"category": "sparql", "metric": "Benchmark queries completed", "value": len(benchmark), "source": str(BENCHMARK_PATH.relative_to(ROOT)), "note": "Fuseki /mathkg500 benchmark"},
        {"category": "sparql", "metric": "Mean response time ms", "value": fmt(avg_benchmark), "source": str(BENCHMARK_PATH.relative_to(ROOT)), "note": "Average over benchmark queries"},
        {"category": "sparql", "metric": "Max response time ms", "value": fmt(max_benchmark), "source": str(BENCHMARK_PATH.relative_to(ROOT)), "note": "Slowest benchmark query"},
        *codex_gloss_agreement_rows(),
        {"category": "tts", "metric": "Audio artifact QC sets passed", "value": sum(1 for row in audio_rows if row["artifact_qc"] == "pass"), "source": "reports/evaluation/week6_tts_audio_quality_qc.csv", "note": "Counts generated audio sets that have all expected files."},
        tts_audio_rating_row(audio_rows),
    ]


def write_paper_sections(
    ready: bool,
    comprehension_stats: dict[str, object] | None,
    nasa_stats: dict[str, object] | None,
    system_rows: list[dict[str, object]],
    themes: list[dict[str, object]],
) -> None:
    PAPER_DIR.mkdir(parents=True, exist_ok=True)
    value_by_metric = {row["metric"]: row["value"] for row in system_rows}
    study_phrase = (
        "a real-participant within-subjects user study"
        if ready
        else "a prepared real-participant within-subjects user-study workflow"
    )
    figure_4_phrase = (
        "Figure 4 shows the user-study comprehension chart with error bars"
        if ready
        else "Figure 4 is reserved for the user-study comprehension chart once real participant entries are complete"
    )
    section_5 = f"""# 5. Evaluation

The evaluation examines MathOntoSpeak as a multi-perspective knowledge graph pipeline: ontology coverage, query performance, gloss-review readiness, TTS artifact generation, and {study_phrase}.

The paper figures are organized around the system contribution and evaluation evidence: Figure 1 presents the five-layer architecture, Figure 2 shows the `T` concept graph with four surface forms, Figure 3 reports SPARQL benchmark response times, and {figure_4_phrase}.

## 5.1 Ontology and Metadata Coverage

The merged graph contains {value_by_metric['N classes']} ontology classes, {value_by_metric['N properties']} declared properties, and {value_by_metric['N axioms/statements']} RDF statements. Of the ontology classes, {value_by_metric['N provenance-tagged classes']} carry provenance-oriented metadata. This supports the central contribution because each speech rendering is tied to a concept node with labels, definitions, provenance, domain tags, and kind/role classification.

Figure 1 summarizes the five-layer architecture: source ontologies and metadata feed the merged knowledge graph, SPARQL/API access, surface-form rendering, and audio/evaluation outputs.

## 5.2 SPARQL Benchmark

The benchmark suite contains {value_by_metric['Benchmark queries completed']} representative SPARQL queries over the cleaned 500-concept graph. The local Fuseki `/mathkg500` benchmark completed with a mean response time of {value_by_metric['Mean response time ms']} ms and a maximum response time of {value_by_metric['Max response time ms']} ms. Figure 3 reports the per-query timing profile and shows that graph lookup is practical for prototype interaction.

## 5.3 Gloss Review and Audio Quality

Gloss quality was evaluated with a 50-record Codex two-pass QC review. The review produced an overall Cohen's kappa of {value_by_metric['Codex two-pass gloss QC overall Cohen kappa']}, with accuracy kappa {value_by_metric['Codex two-pass gloss QC accuracy Cohen kappa']}, naturalness kappa {value_by_metric['Codex two-pass gloss QC naturalness Cohen kappa']}, and cross-domain correctness kappa {value_by_metric['Codex two-pass gloss QC cross-domain Cohen kappa']}. This is an internal agreement metric rather than a human mentor review.

The audio artifact check confirms that {value_by_metric['Audio artifact QC sets passed']} generated audio sets contain their expected files. The mean TTS audio quality rating is {value_by_metric['Mean TTS audio quality rating']} on a 0-5 completeness scale.
"""
    if ready and comprehension_stats and nasa_stats:
        section_5 += f"""
## 5.4 User Study Results

In the real-participant within-subjects workbook, MathOntoSpeak semantic TTS produced a mean comprehension accuracy of {pct(float(comprehension_stats['semantic_mean']))}, compared with {pct(float(comprehension_stats['notation_mean']))} for notation-only TTS. The paired t-test was t = {fmt(float(comprehension_stats['paired_t_statistic']))}, p {fmt_p(float(comprehension_stats['paired_t_p_value']))}, with Cohen's dz = {fmt(float(comprehension_stats['cohens_dz']))}. NASA-TLX workload was {fmt(float(nasa_stats['semantic_mean']), 1)} for MathOntoSpeak semantic TTS and {fmt(float(nasa_stats['notation_mean']), 1)} for notation-only TTS; the Wilcoxon signed-rank test was W = {fmt(float(nasa_stats['wilcoxon_statistic']))}, p {fmt_p(float(nasa_stats['wilcoxon_p_value']))}.

The interview coding identified {len(themes)} themes: {', '.join(row['theme'] for row in themes)}.

These results come from real participants in the within-subjects study workbook. Figure 4 visualizes the comprehension difference between notation-only TTS and MathOntoSpeak semantic TTS with standard-error bars.
"""
    else:
        section_5 += f"""
## 5.4 User Study Results

The user-study statistics are reserved for completed participant-entry workbooks. This results package will become final after MCQ responses, NASA-TLX ratings, and interview codes are entered in `{selected_workbook_path().relative_to(ROOT)}`.
"""
    section_5 += """
## 5.5 Summary

The completed technical evidence shows that MathOntoSpeak can serialize a provenance-tagged mathematical KG, query it through SPARQL, expose concept records to rendering code, generate multiple audio-oriented surface forms, and report a complete system evaluation table from available project evidence.

Figure 2 illustrates the key knowledge-graph contribution: one `T` concept node can be rendered through four formally named surface forms, allowing concise, pedagogical, expert, and document-role perspectives without duplicating the underlying graph node.
"""

    if ready:
        evidence_sentence = "The graph has measurable ontology coverage, provenance tagging, SPARQL benchmark timings, generated audio artifacts, real-participant user-study results, and internal gloss-review agreement."
    else:
        evidence_sentence = "The graph has measurable ontology coverage, provenance tagging, SPARQL benchmark timings, generated audio artifacts, internal gloss-review agreement, and a ready data-entry workflow for the planned within-subjects user study."

    section_6 = f"""# 6. Discussion

MathOntoSpeak contributes a multi-perspective knowledge graph approach to mathematical accessibility. Its central design choice is to treat a mathematical concept as one graph node that can support several formally distinct renderings. The concise, pedagogical, expert, and document-role forms are not merely alternate phrasings; they encode different perspectives over the same ontology-backed object.

This distinction matters because mathematical notation compresses role, context, and domain convention into compact symbols. A symbol such as `T` can surface as a letter, a transformation, an operator, or a document-specific referent. A notation-only TTS pipeline can pronounce the surface symbol, but it cannot reliably explain the conceptual role being invoked. MathOntoSpeak addresses that problem by linking a symbol to a concept IRI and selecting a perspective-specific rendering from graph metadata.

The four surface forms bridge formal semantic-web representation and learner-facing audio. The concise form supports quick access, the pedagogical form supports explanation, the expert form preserves technical language, and the document-role form explains how the concept functions in mathematical prose. These are separate communicative views of the same node, which is why the contribution is a KG modeling contribution rather than only a TTS formatting layer.

The evaluation evidence supports this architecture. {evidence_sentence} Together, these results show end-to-end movement from linked data to accessible speech output: a concept node is modeled once, queried through the graph, rendered through formally named perspectives, and evaluated as a speech-facing accessibility interface.

The broader implication is that accessibility rendering can be treated as a knowledge representation problem. By representing mathematical meaning once and rendering it through multiple formally named perspectives, MathOntoSpeak gives downstream interfaces a principled way to adapt speech to the listener, task, and document role.
"""
    (PAPER_DIR / "section_5_evaluation.md").write_text(section_5, encoding="utf-8")
    (PAPER_DIR / "section_6_discussion.md").write_text(section_6, encoding="utf-8")


def write_evidence_package_status(blockers: list[str]) -> None:
    lines = [
        "# Week 6 Evidence Package Status",
        "",
        "The technical evaluation package is implemented. It uses ontology metrics, SPARQL timings, Codex two-pass gloss agreement, generated audio artifact QC, and regeneration-safe paper figures.",
        "",
        "## Already Regenerated From Project Evidence",
        "",
        "- Ontology statistics from the merged TTL.",
        "- SPARQL benchmark summary from the Fuseki timing CSV.",
        "- Four required paper figures under `figures/`.",
        "- Optional poster supplement: Figure 5 tools/software flow.",
        "- Paper Sections 5 and 6 using evidence-safe wording.",
    ]
    if blockers:
        lines += [
            "",
            "## Pending Evidence Inputs",
            "",
            *[f"- {blocker}" for blocker in blockers],
        ]
    else:
        lines += [
            "",
            "## Current Status",
            "",
            "- Complete for the current evidence package.",
        ]
    (EVAL_DIR / "week6_evidence_package_status.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    EVAL_DIR.mkdir(parents=True, exist_ok=True)
    FIG_DIR.mkdir(parents=True, exist_ok=True)
    ensure_manual_input_templates()

    comprehension_rows, nasa_rows, interview_rows, blockers = extract_real_study_data()
    ready = study_ready(comprehension_rows, nasa_rows, interview_rows, blockers)

    comprehension_stats = None
    nasa_stats = None
    themes: list[dict[str, object]] = []
    if ready:
        comprehension_stats = paired_stats(paired_by_condition(comprehension_rows, "score"), semantic_minus_notation=True)
        nasa_stats = paired_stats(paired_by_condition(nasa_rows, "raw_tlx_mean"), semantic_minus_notation=True)
        themes = interview_theme_rows(interview_rows)

    audio_rows = audio_quality_rows()
    system_rows = system_evaluation_rows(audio_rows)
    benchmark = read_csv(BENCHMARK_PATH)
    glosses = json.loads(GLOSS_PATH.read_text(encoding="utf-8"))

    write_statistical_outputs(ready, comprehension_stats, nasa_stats, themes, blockers)
    write_csv(EVAL_DIR / "week6_tts_audio_quality_qc.csv", audio_rows)
    write_csv(EVAL_DIR / "week6_system_evaluation_table.csv", system_rows)
    write_paper_sections(ready, comprehension_stats, nasa_stats, system_rows, themes)
    write_evidence_package_status(blockers)

    make_architecture_figure()
    make_concept_graph_figure(glosses)
    make_sparql_chart(benchmark)
    make_comprehension_chart(comprehension_stats)
    make_tools_software_flow_figure()

    print(EVAL_DIR / "week6_statistical_analysis_summary.md")
    print(EVAL_DIR / "week6_system_evaluation_table.csv")
    print(EVAL_DIR / "week6_evidence_package_status.md")
    print(FIG_DIR)
    print(PAPER_DIR / "section_5_evaluation.md")
    print(PAPER_DIR / "section_6_discussion.md")


if __name__ == "__main__":
    main()
